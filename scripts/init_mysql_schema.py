import argparse
from pathlib import Path

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[1]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=r"D:\document\up\mysql.xlsx")
    parser.add_argument("--schema", default=str(ROOT / "delivery" / "sql" / "mysql_schema.sql"))
    args = parser.parse_args()

    cfg = read_mysql_excel(Path(args.source))
    sql = Path(args.schema).read_text(encoding="utf-8")
    statements = [stmt.strip() for stmt in sql.split(";") if stmt.strip()]

    conn = pymysql.connect(
        host=cfg["host"],
        port=cfg["port"],
        user=cfg["user"],
        password=cfg["password"],
        database=cfg["database"],
        charset="utf8mb4",
        autocommit=True,
    )
    try:
        with conn.cursor() as cursor:
            for statement in statements:
                cursor.execute(statement)
        print("MySQL schema initialized: adapter core, Apifox config, DingTalk config, Yunxiao config")
    finally:
        conn.close()


def read_mysql_excel(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    values = {headers[i]: ws.cell(row=2, column=i + 1).value for i in range(len(headers))}
    return {
        "host": str(values["地址"]).strip(),
        "port": 3306,
        "database": str(values["数据库"]).strip(),
        "user": str(values["账号"]).strip(),
        "password": str(values["密码"]).strip(),
    }


if __name__ == "__main__":
    main()
